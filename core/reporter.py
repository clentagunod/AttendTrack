"""core/reporter.py — Generates formatted Excel summary reports."""

import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.chart.series import DataPoint
from datetime import datetime
from pathlib import Path


# ── Style helpers ──────────────────────────────────────────────────────────────

def _hex(h: str):
    return h.lstrip("#")


def _font(bold=False, size=11, color="000000", italic=False):
    return Font(name="Calibri", bold=bold, size=size,
                color=_hex(color), italic=italic)


def _fill(color: str):
    return PatternFill("solid", fgColor=_hex(color))


def _border(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)


def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _apply_header_row(ws, row_num: int, cols: list,
                      bg: str, fg: str = "FFFFFF", size=11):
    for i, val in enumerate(cols, 1):
        c = ws.cell(row=row_num, column=i, value=val)
        c.font   = _font(bold=True, size=size, color=fg)
        c.fill   = _fill(bg)
        c.border = _border()
        c.alignment = _align("center")


def _apply_data_row(ws, row_num: int, values: list,
                    bg_even: str = "F0F4F8", bg_odd: str = "FFFFFF",
                    is_even: bool = True):
    bg = bg_even if is_even else bg_odd
    for i, val in enumerate(values, 1):
        c = ws.cell(row=row_num, column=i, value=val)
        c.font   = _font(size=10)
        c.fill   = _fill(bg)
        c.border = _border("hair")
        c.alignment = _align("center")


def _auto_width(ws, min_w=10, max_w=40):
    for col in ws.columns:
        best = min_w
        for cell in col:
            try:
                best = max(best, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(best + 3, max_w)


# ── Main report builder ────────────────────────────────────────────────────────

def generate_report(
    raw_df: pd.DataFrame,
    computed_df: pd.DataFrame,
    emp_summary: pd.DataFrame,
    dept_summary: pd.DataFrame,
    kpis: dict,
    settings: dict,
    output_path: str | Path,
) -> Path:
    """
    Write a multi-sheet formatted Excel report.
    Returns the path written.
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    rcfg     = settings["report"]
    HDR      = rcfg.get("header_color", "2C5F8A")
    ACCENT   = rcfg.get("accent_color", "1E3A5F")
    COMPANY  = rcfg.get("company_name", "My Company")
    now_str  = datetime.now().strftime("%Y-%m-%d %H:%M")

    wb = Workbook()

    # ── Sheet 1: Dashboard ────────────────────────────────────────────────────
    ws_dash = wb.active
    ws_dash.title = "Dashboard"
    ws_dash.sheet_view.showGridLines = False

    # Title block
    ws_dash.merge_cells("A1:H1")
    t = ws_dash["A1"]
    t.value     = f"{COMPANY} — Attendance Report"
    t.font      = _font(bold=True, size=20, color="FFFFFF")
    t.fill      = _fill(ACCENT)
    t.alignment = _align("center")
    ws_dash.row_dimensions[1].height = 40

    ws_dash.merge_cells("A2:H2")
    sub = ws_dash["A2"]
    sub.value     = f"Generated: {now_str}"
    sub.font      = _font(size=10, color="FFFFFF", italic=True)
    sub.fill      = _fill(HDR)
    sub.alignment = _align("center")
    ws_dash.row_dimensions[2].height = 20

    # KPI tiles (row 4+)
    ws_dash.row_dimensions[3].height = 12
    kpi_items = [
        ("Total Employees",  kpis.get("total_employees", 0),   "1A73E8"),
        ("Attendance Rate",  f"{kpis.get('attendance_rate', 0)}%", "34A853"),
        ("Punctuality Rate", f"{kpis.get('punctuality_rate', 0)}%", "FBBC05"),
        ("Total Absent",     kpis.get("total_absent", 0),      "EA4335"),
        ("Total Late",       kpis.get("total_late", 0),        "FF6D00"),
        ("Avg Hours/Day",    kpis.get("avg_hours", 0),         "7B1FA2"),
        ("Total Overtime",   kpis.get("total_overtime", 0),    "00838F"),
        ("Records Analyzed", kpis.get("total_records", 0),     "37474F"),
    ]

    col_pairs = [(1, 2), (3, 4), (5, 6), (7, 8),
                 (1, 2), (3, 4), (5, 6), (7, 8)]
    row_offsets = [4, 4, 4, 4, 8, 8, 8, 8]

    for idx, (label, value, color) in enumerate(kpi_items):
        r = row_offsets[idx]
        c1, c2 = col_pairs[idx]

        ws_dash.merge_cells(
            start_row=r, start_column=c1, end_row=r, end_column=c2
        )
        ws_dash.merge_cells(
            start_row=r+1, start_column=c1, end_row=r+1, end_column=c2
        )
        ws_dash.merge_cells(
            start_row=r+2, start_column=c1, end_row=r+2, end_column=c2
        )

        for dr in range(3):
            for dc in range(c2 - c1 + 1):
                cell = ws_dash.cell(row=r+dr, column=c1+dc)
                cell.fill = _fill(color)

        lbl_cell = ws_dash.cell(row=r, column=c1, value=label.upper())
        lbl_cell.font      = _font(bold=True, size=8, color="FFFFFF")
        lbl_cell.alignment = _align("center")

        val_cell = ws_dash.cell(row=r+1, column=c1, value=value)
        val_cell.font      = _font(bold=True, size=22, color="FFFFFF")
        val_cell.alignment = _align("center")

        ws_dash.row_dimensions[r].height   = 18
        ws_dash.row_dimensions[r+1].height = 36
        ws_dash.row_dimensions[r+2].height = 8

    for c in range(1, 9):
        ws_dash.column_dimensions[get_column_letter(c)].width = 18

    # ── Sheet 2: Employee Summary ─────────────────────────────────────────────
    ws_emp = wb.create_sheet("Employee Summary")
    ws_emp.sheet_view.showGridLines = False

    ws_emp.merge_cells("A1:J1")
    hdr = ws_emp["A1"]
    hdr.value     = "Employee Attendance Summary"
    hdr.font      = _font(bold=True, size=14, color="FFFFFF")
    hdr.fill      = _fill(HDR)
    hdr.alignment = _align("center")
    ws_emp.row_dimensions[1].height = 30

    if not emp_summary.empty:
        cols = list(emp_summary.columns)
        _apply_header_row(ws_emp, 2, cols, HDR, size=10)
        for i, (_, row) in enumerate(emp_summary.iterrows()):
            _apply_data_row(ws_emp, 3 + i, list(row), is_even=(i % 2 == 0))

    _auto_width(ws_emp)

    # ── Sheet 3: Department Summary ───────────────────────────────────────────
    ws_dept = wb.create_sheet("Department Summary")
    ws_dept.sheet_view.showGridLines = False

    ws_dept.merge_cells("A1:G1")
    hdr2 = ws_dept["A1"]
    hdr2.value     = "Department Attendance Summary"
    hdr2.font      = _font(bold=True, size=14, color="FFFFFF")
    hdr2.fill      = _fill(HDR)
    hdr2.alignment = _align("center")
    ws_dept.row_dimensions[1].height = 30

    if not dept_summary.empty:
        cols = list(dept_summary.columns)
        _apply_header_row(ws_dept, 2, cols, HDR, size=10)
        for i, (_, row) in enumerate(dept_summary.iterrows()):
            _apply_data_row(ws_dept, 3 + i, list(row), is_even=(i % 2 == 0))

    _auto_width(ws_dept)

    # ── Sheet 4: Raw Detail ───────────────────────────────────────────────────
    ws_raw = wb.create_sheet("Attendance Detail")
    ws_raw.sheet_view.showGridLines = False

    detail_cols = [
        "employee_name", "department", "date", "time_in", "time_out",
        "status_normalized", "hours_worked", "is_late", "late_minutes", "overtime_hours"
    ]
    show_cols   = [c for c in detail_cols if c in computed_df.columns]
    header_nice = {
        "employee_name":    "Employee",
        "department":       "Department",
        "date":             "Date",
        "time_in":          "Time In",
        "time_out":         "Time Out",
        "status_normalized":"Status",
        "hours_worked":     "Hours Worked",
        "is_late":          "Late?",
        "late_minutes":     "Late (min)",
        "overtime_hours":   "Overtime Hrs",
    }

    ws_raw.merge_cells("A1:J1")
    hdr3 = ws_raw["A1"]
    hdr3.value     = "Full Attendance Detail"
    hdr3.font      = _font(bold=True, size=14, color="FFFFFF")
    hdr3.fill      = _fill(ACCENT)
    hdr3.alignment = _align("center")
    ws_raw.row_dimensions[1].height = 28

    _apply_header_row(ws_raw, 2, [header_nice.get(c, c) for c in show_cols], HDR, size=10)

    for i, (_, row) in enumerate(computed_df[show_cols].iterrows()):
        vals = []
        for col in show_cols:
            v = row[col]
            if col in ("time_in", "time_out") and v is not None:
                try:
                    v = v.strftime("%H:%M")
                except Exception:
                    v = str(v)
            elif col == "is_late":
                v = "Yes" if v else "No"
            elif col == "date" and v is not None:
                try:
                    v = v.strftime("%Y-%m-%d")
                except Exception:
                    v = str(v)
            vals.append(v)

        _apply_data_row(ws_raw, 3 + i, vals, is_even=(i % 2 == 0))

        # Highlight late rows in light red
        if row.get("is_late", False):
            for ci in range(1, len(show_cols) + 1):
                ws_raw.cell(3 + i, ci).fill = _fill("FFE0E0")

    _auto_width(ws_raw)

    wb.save(str(output_path))
    return output_path


def generate_sample_template(path: str | Path, settings: dict) -> Path:
    """Generate a sample attendance Excel template."""
    path = Path(path)
    col_map = settings["excel"]["expected_columns"]
    cols    = list(col_map.values())

    sample_data = [
        ["E001", "Juan Santos",    "HR",          "2025-06-01", "08:05", "17:10", "Present"],
        ["E001", "Juan Santos",    "HR",          "2025-06-02", "08:45", "17:00", "Present"],
        ["E002", "Maria Reyes",    "Finance",     "2025-06-01", "08:00", "17:00", "Present"],
        ["E002", "Maria Reyes",    "Finance",     "2025-06-02", "",      "",      "Absent"],
        ["E003", "Carlo Mendoza",  "IT",          "2025-06-01", "07:55", "17:30", "Present"],
        ["E003", "Carlo Mendoza",  "IT",          "2025-06-02", "09:00", "13:00", "Half Day"],
        ["E004", "Ana Dela Cruz",  "Operations",  "2025-06-01", "08:10", "17:05", "Present"],
        ["E004", "Ana Dela Cruz",  "Operations",  "2025-06-02", "08:00", "17:00", "Present"],
        ["E005", "Renz Villanueva","HR",          "2025-06-01", "",      "",      "Leave"],
        ["E005", "Renz Villanueva","HR",          "2025-06-02", "08:20", "17:00", "Present"],
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    HDR = "2C5F8A"
    _apply_header_row(ws, 1, cols, HDR)
    ws.row_dimensions[1].height = 22

    for i, row in enumerate(sample_data):
        _apply_data_row(ws, 2 + i, row, is_even=(i % 2 == 0))

    _auto_width(ws)
    wb.save(str(path))
    return path